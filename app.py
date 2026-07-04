#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
营销人员信息管理系统 - 后端应用
基于 Flask + SQLite，支持多用户登录、联系人管理、跟进记录、数据统计等功能。
"""

import os
import sqlite3
import json
import hashlib
import secrets
import time
from datetime import datetime, date
from functools import wraps
from flask import (Flask, request, session, redirect, url_for, render_template,
                   jsonify, g, send_file, abort, flash)
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)

# 安全密钥：优先从环境变量读取，否则自动生成（云部署务必设置环境变量 SECRET_KEY）
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# 数据库路径：支持环境变量自定义（便于云服务器挂载数据盘持久化）
_db_path = os.environ.get('DATABASE_PATH', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'marketing.db'))
app.config['DATABASE'] = _db_path

# 运行模式配置
app.config['DEBUG'] = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
app.config['PORT'] = int(os.environ.get('PORT', 5000))

# ============================================================
# 数据库连接管理
# ============================================================
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db


@app.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    """初始化数据库表结构"""
    db = sqlite3.connect(app.config['DATABASE'])
    db.row_factory = sqlite3.Row
    db.executescript('''
        -- 用户表
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            name TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'staff',  -- admin / staff
            access_level TEXT DEFAULT 'L1',       -- L1=仅本人 L2=同组 L3=全部
            team TEXT DEFAULT '',                  -- 所属小组
            phone TEXT,
            email TEXT,
            status INTEGER DEFAULT 1,  -- 1=启用 0=禁用
            created_at TEXT DEFAULT (datetime('localtime')),
            last_login TEXT
        );

        -- 联系人表
        CREATE TABLE IF NOT EXISTS contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,        -- 大类
            subcategory TEXT DEFAULT '',   -- 子类
            unit_name TEXT NOT NULL,       -- 单位名称
            department TEXT DEFAULT '',    -- 部门
            position TEXT DEFAULT '',      -- 职务
            name TEXT DEFAULT '',          -- 姓名
            phone TEXT DEFAULT '',         -- 联系方式
            main_project TEXT DEFAULT '',  -- 主要跟踪项目/目标
            office_address TEXT DEFAULT '',-- 办公地址
            delivery_address TEXT DEFAULT '',-- 快递地址
            maintenance_level TEXT DEFAULT 'B', -- 维系等级 A/B/C/D
            maintenance_person TEXT DEFAULT '', -- 维系人
            remark TEXT DEFAULT '',        -- 备注
            status TEXT DEFAULT '活跃',     -- 活跃/潜在/沉睡
            owner_id INTEGER,              -- 归属用户ID
            created_by INTEGER,
            created_at TEXT DEFAULT (datetime('localtime')),
            updated_at TEXT DEFAULT (datetime('localtime')),
            FOREIGN KEY (created_by) REFERENCES users(id)
        );

        -- 跟进记录表
        CREATE TABLE IF NOT EXISTS followups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contact_id INTEGER NOT NULL,
            follow_date TEXT NOT NULL,       -- 跟进日期
            follow_type TEXT DEFAULT '电话',  -- 电话/拜访/微信/邮件/会议
            content TEXT NOT NULL,            -- 跟进内容
            next_follow_date TEXT,            -- 下次跟进日期
            created_by INTEGER,
            created_at TEXT DEFAULT (datetime('localtime')),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE,
            FOREIGN KEY (created_by) REFERENCES users(id)
        );

        -- 操作日志表
        CREATE TABLE IF NOT EXISTS operation_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT,
            target TEXT,
            detail TEXT,
            ip TEXT,
            created_at TEXT DEFAULT (datetime('localtime')),
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
    ''')
    db.commit()

    # 创建默认管理员
    admin = db.execute('SELECT id FROM users WHERE username = ?', ('admin',)).fetchone()
    if not admin:
        db.execute(
            'INSERT INTO users (username, password_hash, name, role, access_level, team, phone) VALUES (?, ?, ?, ?, ?, ?, ?)',
            ('admin', generate_password_hash('admin123'), '系统管理员', 'admin', 'L3', '管理组', '13800000000')
        )
        # 创建一个示例营销人员账号
        db.execute(
            'INSERT OR IGNORE INTO users (username, password_hash, name, role, access_level, team, phone) VALUES (?, ?, ?, ?, ?, ?, ?)',
            ('sales01', generate_password_hash('sales123'), '营销员-张三', 'staff', 'L1', '一组', '13900000001')
        )
        db.commit()
        print('[初始化] 默认管理员账号: admin / admin123')
        print('[初始化] 示例营销员账号: sales01 / sales123')

    db.close()


# ============================================================
# 认证与权限
# ============================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'ok': False, 'msg': '请先登录'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated


def current_user():
    if 'user_id' in session:
        db = get_db()
        return db.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    return None


def log_action(action, target='', detail=''):
    """记录操作日志"""
    if 'user_id' in session:
        db = get_db()
        db.execute(
            'INSERT INTO operation_logs (user_id, action, target, detail, ip) VALUES (?, ?, ?, ?, ?)',
            (session['user_id'], action, target, detail, request.remote_addr or '')
        )
        db.commit()


# ============================================================
# 常量定义
# ============================================================
CATEGORIES = {
    '业主单位': ['政府部门', '主管单位', '投资方', '学校'],
    '设计及咨询单位': [],
    '施工单位及监理': [],
    '供应商': [],
    '运营方': [],
}

MAINTENANCE_LEVELS = {
    'A': {'label': 'A级-核心', 'color': '#e74c3c', 'desc': '核心客户，高频维系'},
    'B': {'label': 'B级-重要', 'color': '#f39c12', 'desc': '重要客户，定期维系'},
    'C': {'label': 'C级-一般', 'color': '#3498db', 'desc': '一般客户，适时维系'},
    'D': {'label': 'D级-潜在', 'color': '#95a5a6', 'desc': '潜在客户，观望维系'},
}

FOLLOW_TYPES = ['电话', '拜访', '微信', '邮件', '会议', '其他']

CONTACT_STATUSES = ['活跃', '潜在', '沉睡']

# 权限等级定义
ACCESS_LEVELS = {
    'L1': {'label': 'L1-仅本人', 'color': '#e74c3c', 'desc': '仅查看自己的联系人'},
    'L2': {'label': 'L2-同组', 'color': '#f39c12', 'desc': '查看同组所有联系人'},
    'L3': {'label': 'L3-全部', 'color': '#00B884', 'desc': '查看所有联系人'},
}

# ============================================================
# 权限辅助函数
# ============================================================
def apply_access_filter(query, params, db):
    """
    根据当前登录用户的权限等级，在SQL查询中追加数据过滤条件。
    返回修改后的 (query, params)。
    过滤条件会插入到 ORDER BY / GROUP BY / LIMIT 之前。
    """
    # 管理员始终看全部
    if session.get('role') == 'admin':
        return query, params

    user_id = session.get('user_id')
    level = session.get('access_level', 'L1')
    team = session.get('team', '')

    # 构造过滤条件（直接拼SQL，不用参数化避免 params 顺序问题）
    filter_clause = ''
    if level == 'L1':
        filter_clause = f' AND owner_id = {user_id}'
    elif level == 'L2' and team:
        team_members = db.execute(
            'SELECT id FROM users WHERE team = ? AND status = 1', (team,)
        ).fetchall()
        member_ids = [str(m['id']) for m in team_members]
        if member_ids:
            filter_clause = f' AND owner_id IN ({",".join(member_ids)})'
        else:
            filter_clause = ' AND 1=0'
    # L3: 不追加过滤

    if not filter_clause:
        return query, params

    import re
    # 如果 query 中完全没有 WHERE，需要插入 WHERE 1=1
    if ' WHERE ' not in query.upper():
        # 在第一个 ORDER BY / GROUP BY / LIMIT 之前插入 WHERE 1=1
        # 如果没有这些关键字，在末尾添加
        match = re.search(r'\b(ORDER\s+BY|GROUP\s+BY|LIMIT)\b', query, re.IGNORECASE)
        if match:
            pos = match.start()
            query = query[:pos] + ' WHERE 1=1 ' + query[pos:]
        else:
            query += ' WHERE 1=1'

    # 将过滤条件插入到 ORDER BY / GROUP BY / LIMIT 之前
    match = re.search(r'\b(ORDER\s+BY|GROUP\s+BY|LIMIT)\b', query, re.IGNORECASE)
    if match:
        pos = match.start()
        query = query[:pos] + filter_clause + ' ' + query[pos:]
    else:
        query += filter_clause

    return query, params


# ============================================================
# 路由 - 认证
# ============================================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember')

        db = get_db()
        user = db.execute('SELECT * FROM users WHERE username = ? AND status = 1', (username,)).fetchone()
        if user and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['name'] = user['name']
            session['role'] = user['role']
            session['access_level'] = user['access_level'] or 'L1'
            session['team'] = user['team'] or ''
            session.permanent = bool(remember)
            db.execute('UPDATE users SET last_login = datetime("localtime") WHERE id = ?', (user['id'],))
            db.commit()
            log_action('登录', '系统', f'用户 {username} 登录')
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error='用户名或密码错误，或账号已被禁用')
    return render_template('login.html')


@app.route('/logout')
def logout():
    log_action('登出', '系统', f'用户 {session.get("username", "")} 登出')
    session.clear()
    return redirect(url_for('login'))


# ============================================================
# 路由 - 页面
# ============================================================
@app.route('/')
@login_required
def dashboard():
    """仪表盘 - 数据统计概览"""
    db = get_db()
    # 总联系人数（按权限过滤）
    q, p = apply_access_filter('SELECT COUNT(*) as c FROM contacts', [], db)
    total = db.execute(q, p).fetchone()['c']
    # 按大类统计
    q, p = apply_access_filter('SELECT category, COUNT(*) as c FROM contacts GROUP BY category ORDER BY c DESC', [], db)
    by_category = db.execute(q, p).fetchall()
    # 按维系等级统计
    q, p = apply_access_filter('SELECT maintenance_level, COUNT(*) as c FROM contacts GROUP BY maintenance_level', [], db)
    by_level = db.execute(q, p).fetchall()
    # 按状态统计
    q, p = apply_access_filter('SELECT status, COUNT(*) as c FROM contacts GROUP BY status', [], db)
    by_status = db.execute(q, p).fetchall()
    # 最近跟进记录
    q, p = apply_access_filter('''
        SELECT f.*, c.unit_name, c.name as contact_name, u.name as operator
        FROM followups f
        JOIN contacts c ON f.contact_id = c.id
        LEFT JOIN users u ON f.created_by = u.id''', [], db)
    q += ' ORDER BY f.follow_date DESC LIMIT 10'
    recent_followups = db.execute(q, p).fetchall()
    # 待跟进（下次跟进日期在7天内）
    today = date.today().isoformat()
    q, p = apply_access_filter('''
        SELECT f.*, c.unit_name, c.name as contact_name, c.maintenance_level
        FROM followups f
        JOIN contacts c ON f.contact_id = c.id
        WHERE f.next_follow_date IS NOT NULL AND f.next_follow_date >= ?''', [today], db)
    q += ' ORDER BY f.next_follow_date ASC LIMIT 15'
    upcoming = db.execute(q, p).fetchall()
    # 本月新增
    this_month = datetime.now().strftime('%Y-%m')
    q, p = apply_access_filter(
        "SELECT COUNT(*) as c FROM contacts WHERE created_at LIKE ?", [f'{this_month}%'], db)
    month_new = db.execute(q, p).fetchone()['c']
    # 营销人员统计
    staff_stats = db.execute('''
        SELECT u.name, COUNT(c.id) as count
        FROM users u
        LEFT JOIN contacts c ON c.maintenance_person = u.name
        WHERE u.role = 'staff'
        GROUP BY u.id ORDER BY count DESC
    ''').fetchall()

    return render_template('dashboard.html',
                           total=total, by_category=by_category, by_level=by_level,
                           by_status=by_status, recent_followups=recent_followups,
                           upcoming=upcoming, month_new=month_new, staff_stats=staff_stats,
                           levels=MAINTENANCE_LEVELS)


@app.route('/contacts')
@login_required
def contacts():
    """联系人列表"""
    db = get_db()
    # 获取筛选参数
    category = request.args.get('category', '')
    subcategory = request.args.get('subcategory', '')
    level = request.args.get('level', '')
    status = request.args.get('status', '')
    keyword = request.args.get('keyword', '').strip()
    maintenance_person = request.args.get('person', '')

    query = 'SELECT * FROM contacts WHERE 1=1'
    params = []
    if category:
        query += ' AND category = ?'
        params.append(category)
    if subcategory:
        query += ' AND subcategory = ?'
        params.append(subcategory)
    if level:
        query += ' AND maintenance_level = ?'
        params.append(level)
    if status:
        query += ' AND status = ?'
        params.append(status)
    if maintenance_person:
        query += ' AND maintenance_person = ?'
        params.append(maintenance_person)
    if keyword:
        query += ''' AND (unit_name LIKE ? OR name LIKE ? OR department LIKE ?
                     OR position LIKE ? OR main_project LIKE ? OR phone LIKE ?)'''
        kw = f'%{keyword}%'
        params.extend([kw] * 6)
    query += ' ORDER BY category, subcategory, maintenance_level, unit_name'

    # 权限过滤
    query, params = apply_access_filter(query, params, db)
    rows = db.execute(query, params).fetchall()

    # 获取所有维系人列表
    persons = db.execute(
        "SELECT DISTINCT maintenance_person FROM contacts WHERE maintenance_person != '' ORDER BY maintenance_person"
    ).fetchall()
    # 获取所有营销人员
    staff_list = db.execute("SELECT name FROM users WHERE role='staff' AND status=1 ORDER BY name").fetchall()

    return render_template('contacts.html', contacts=rows,
                           categories=CATEGORIES, levels=MAINTENANCE_LEVELS,
                           statuses=CONTACT_STATUSES, persons=persons, staff_list=staff_list,
                           filters={'category': category, 'subcategory': subcategory,
                                    'level': level, 'status': status, 'keyword': keyword,
                                    'person': maintenance_person})


@app.route('/contacts/new', methods=['GET', 'POST'])
@login_required
def contact_new():
    if request.method == 'POST':
        return save_contact(None)
    db = get_db()
    staff_list = db.execute("SELECT name FROM users WHERE status=1 ORDER BY name").fetchall()
    return render_template('contact_form.html', contact=None, categories=CATEGORIES,
                           levels=MAINTENANCE_LEVELS, statuses=CONTACT_STATUSES,
                           staff_list=staff_list, follow_types=FOLLOW_TYPES)


@app.route('/contacts/<int:cid>/edit', methods=['GET', 'POST'])
@login_required
def contact_edit(cid):
    db = get_db()
    contact = db.execute('SELECT * FROM contacts WHERE id = ?', (cid,)).fetchone()
    if not contact:
        abort(404)
    # 权限校验：非管理员只能编辑自己的联系人
    if session.get('role') != 'admin' and contact['owner_id'] != session.get('user_id'):
        flash('您没有权限编辑此联系人', 'error')
        return redirect(url_for('contacts'))
    if request.method == 'POST':
        return save_contact(cid)
    staff_list = db.execute("SELECT name FROM users WHERE status=1 ORDER BY name").fetchall()
    # 获取该联系人的跟进记录
    followups = db.execute('''
        SELECT f.*, u.name as operator FROM followups f
        LEFT JOIN users u ON f.created_by = u.id
        WHERE f.contact_id = ? ORDER BY f.follow_date DESC
    ''', (cid,)).fetchall()
    return render_template('contact_form.html', contact=contact, categories=CATEGORIES,
                           levels=MAINTENANCE_LEVELS, statuses=CONTACT_STATUSES,
                           staff_list=staff_list, follow_types=FOLLOW_TYPES,
                           followups=followups)


def save_contact(cid):
    """保存联系人（新增或编辑）"""
    db = get_db()
    data = {
        'category': request.form.get('category', '').strip(),
        'subcategory': request.form.get('subcategory', '').strip(),
        'unit_name': request.form.get('unit_name', '').strip(),
        'department': request.form.get('department', '').strip(),
        'position': request.form.get('position', '').strip(),
        'name': request.form.get('name', '').strip(),
        'phone': request.form.get('phone', '').strip(),
        'main_project': request.form.get('main_project', '').strip(),
        'office_address': request.form.get('office_address', '').strip(),
        'delivery_address': request.form.get('delivery_address', '').strip(),
        'maintenance_level': request.form.get('maintenance_level', 'B').strip(),
        'maintenance_person': request.form.get('maintenance_person', '').strip(),
        'remark': request.form.get('remark', '').strip(),
        'status': request.form.get('status', '活跃').strip(),
    }
    if not data['unit_name']:
        flash('单位名称不能为空', 'error')
        return redirect(request.url)

    if cid:
        db.execute('''UPDATE contacts SET
            category=?, subcategory=?, unit_name=?, department=?, position=?, name=?,
            phone=?, main_project=?, office_address=?, delivery_address=?,
            maintenance_level=?, maintenance_person=?, remark=?, status=?,
            updated_at=datetime('localtime') WHERE id=?''',
            (*data.values(), cid))
        db.commit()
        log_action('编辑联系人', data['unit_name'], f'修改联系人ID={cid}')
        # 处理跟进记录
        follow_type = request.form.get('fu_type', '').strip()
        follow_content = request.form.get('fu_content', '').strip()
        if follow_content:
            follow_date = request.form.get('fu_date', date.today().isoformat())
            next_date = request.form.get('fu_next', '').strip() or None
            db.execute('''INSERT INTO followups
                (contact_id, follow_date, follow_type, content, next_follow_date, created_by)
                VALUES (?, ?, ?, ?, ?, ?)''',
                (cid, follow_date, follow_type, follow_content, next_date, session['user_id']))
            db.commit()
        flash('联系人信息已更新', 'success')
    else:
        db.execute('''INSERT INTO contacts
            (category, subcategory, unit_name, department, position, name, phone,
             main_project, office_address, delivery_address, maintenance_level,
             maintenance_person, remark, status, owner_id, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (*data.values(), session['user_id'], session['user_id']))
        db.commit()
        new_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
        log_action('新增联系人', data['unit_name'], f'新增联系人ID={new_id}')
        flash('联系人已添加', 'success')
    return redirect(url_for('contacts'))


@app.route('/contacts/<int:cid>/delete', methods=['POST'])
@login_required
def contact_delete(cid):
    db = get_db()
    contact = db.execute('SELECT unit_name, owner_id FROM contacts WHERE id = ?', (cid,)).fetchone()
    if contact:
        if session.get('role') != 'admin' and contact['owner_id'] != session.get('user_id'):
            flash('您没有权限删除此联系人', 'error')
            return redirect(url_for('contacts'))
        db.execute('DELETE FROM contacts WHERE id = ?', (cid,))
        db.commit()
        log_action('删除联系人', contact['unit_name'], f'删除联系人ID={cid}')
        flash('联系人已删除', 'success')
    return redirect(url_for('contacts'))


# ============================================================
# 跟进记录
# ============================================================
@app.route('/followups')
@login_required
def followups():
    """跟进记录列表"""
    db = get_db()
    keyword = request.args.get('keyword', '').strip()
    follow_type = request.args.get('type', '')

    query = '''SELECT f.*, c.unit_name, c.name as contact_name, c.maintenance_level,
               u.name as operator FROM followups f
               JOIN contacts c ON f.contact_id = c.id
               LEFT JOIN users u ON f.created_by = u.id WHERE 1=1'''
    params = []
    if keyword:
        query += ' AND (c.unit_name LIKE ? OR f.content LIKE ? OR c.name LIKE ?)'
        kw = f'%{keyword}%'
        params.extend([kw] * 3)
    if follow_type:
        query += ' AND f.follow_type = ?'
        params.append(follow_type)
    # 权限过滤
    query, params = apply_access_filter(query, params, db)
    query += ' ORDER BY f.follow_date DESC LIMIT 200'

    rows = db.execute(query, params).fetchall()
    return render_template('followups.html', followups=rows, follow_types=FOLLOW_TYPES,
                           keyword=keyword, cur_type=follow_type)


@app.route('/api/followups/<int:fid>/delete', methods=['POST'])
@login_required
def followup_delete(fid):
    db = get_db()
    fu = db.execute('SELECT contact_id FROM followups WHERE id = ?', (fid,)).fetchone()
    if fu:
        db.execute('DELETE FROM followups WHERE id = ?', (fid,))
        db.commit()
        log_action('删除跟进记录', '', f'删除跟进记录ID={fid}')
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'msg': '记录不存在'}), 404


# ============================================================
# 用户管理（管理员）
# ============================================================
@app.route('/users')
@admin_required
def users():
    db = get_db()
    rows = db.execute('''SELECT u.*, (SELECT COUNT(*) FROM contacts WHERE created_by = u.id) as contact_count
        FROM users u ORDER BY u.role DESC, u.created_at''').fetchall()
    user_list = [dict(r) for r in rows]
    return render_template('users.html', users=user_list)


@app.route('/users/new', methods=['POST'])
@admin_required
def user_new():
    db = get_db()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    name = request.form.get('name', '').strip()
    role = request.form.get('role', 'staff')
    access_level = request.form.get('access_level', 'L1').strip()
    team = request.form.get('team', '').strip()
    phone = request.form.get('phone', '').strip()
    email = request.form.get('email', '').strip()

    if not username or not password or not name:
        flash('用户名、密码和姓名为必填项', 'error')
        return redirect(url_for('users'))
    existing = db.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
    if existing:
        flash('用户名已存在', 'error')
        return redirect(url_for('users'))

    db.execute(
        'INSERT INTO users (username, password_hash, name, role, access_level, team, phone, email) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
        (username, generate_password_hash(password), name, role, access_level, team, phone, email)
    )
    db.commit()
    log_action('新增用户', username, f'新增{role}用户: {name} 等级:{access_level}')
    flash(f'用户 {name} 已添加', 'success')
    return redirect(url_for('users'))


@app.route('/users/<int:uid>/edit', methods=['POST'])
@admin_required
def user_edit(uid):
    db = get_db()
    name = request.form.get('name', '').strip()
    role = request.form.get('role', 'staff')
    access_level = request.form.get('access_level', 'L1').strip()
    team = request.form.get('team', '').strip()
    phone = request.form.get('phone', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    status = int(request.form.get('status', 1))

    if password:
        db.execute('''UPDATE users SET name=?, role=?, access_level=?, team=?, phone=?, email=?, status=?,
                      password_hash=? WHERE id=?''',
                   (name, role, access_level, team, phone, email, status, generate_password_hash(password), uid))
    else:
        db.execute('UPDATE users SET name=?, role=?, access_level=?, team=?, phone=?, email=?, status=? WHERE id=?',
                   (name, role, access_level, team, phone, email, status, uid))
    db.commit()
    log_action('编辑用户', name, f'修改用户ID={uid}')
    flash('用户信息已更新', 'success')
    return redirect(url_for('users'))


@app.route('/users/<int:uid>/delete', methods=['POST'])
@admin_required
def user_delete(uid):
    if uid == session.get('user_id'):
        flash('不能删除当前登录的管理员账号', 'error')
        return redirect(url_for('users'))
    db = get_db()
    user = db.execute('SELECT name FROM users WHERE id = ?', (uid,)).fetchone()
    if user:
        # 先清理关联数据：跟进记录、联系人
        db.execute('DELETE FROM followups WHERE created_by = ?', (uid,))
        db.execute('DELETE FROM followups WHERE contact_id IN (SELECT id FROM contacts WHERE owner_id = ?)', (uid,))
        db.execute('DELETE FROM contacts WHERE owner_id = ?', (uid,))
        db.execute('DELETE FROM contacts WHERE created_by = ?', (uid,))
        db.execute('DELETE FROM operation_logs WHERE user_id = ?', (uid,))
        # 最后删除用户
        db.execute('DELETE FROM users WHERE id = ?', (uid,))
        db.commit()
        log_action('删除用户', user['name'], f'删除用户ID={uid}')
        flash(f'用户 {user["name"]} 及其关联数据已删除', 'success')
    return redirect(url_for('users'))


# ============================================================
# 个人资料 / 修改密码
# ============================================================
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    db = get_db()
    user = current_user()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'update_info':
            name = request.form.get('name', '').strip()
            phone = request.form.get('phone', '').strip()
            email = request.form.get('email', '').strip()
            db.execute('UPDATE users SET name=?, phone=?, email=? WHERE id=?',
                       (name, phone, email, session['user_id']))
            db.commit()
            session['name'] = name
            log_action('修改个人资料', '', f'更新个人信息')
            flash('个人信息已更新', 'success')
        elif action == 'change_password':
            old_pwd = request.form.get('old_password', '')
            new_pwd = request.form.get('new_password', '')
            confirm_pwd = request.form.get('confirm_password', '')
            if not check_password_hash(user['password_hash'], old_pwd):
                flash('原密码错误', 'error')
            elif len(new_pwd) < 6:
                flash('新密码长度不能少于6位', 'error')
            elif new_pwd != confirm_pwd:
                flash('两次输入的新密码不一致', 'error')
            else:
                db.execute('UPDATE users SET password_hash=? WHERE id=?',
                           (generate_password_hash(new_pwd), session['user_id']))
                db.commit()
                log_action('修改密码', '', '用户修改了密码')
                flash('密码修改成功', 'success')
        return redirect(url_for('profile'))
    return render_template('profile.html', user=user)


# ============================================================
# 数据导出
# ============================================================
@app.route('/export')
@login_required
def export_excel():
    """导出联系人数据为Excel（含大类、子类、跟进记录）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    db = get_db()
    q, p = apply_access_filter('SELECT * FROM contacts ORDER BY category, subcategory, id', [], db)
    contacts = db.execute(q, p).fetchall()

    # 获取所有contact_id用于跟进记录过滤
    contact_ids = [c['id'] for c in contacts]

    wb = openpyxl.Workbook()

    # === Sheet1: 联系人数据 ===
    ws = wb.active
    ws.title = '营销人员信息统计表'

    headers = ['序号', '大类', '子类', '单位名称', '部门', '职务', '姓名',
               '联系方式', '主要跟踪项目/目标', '办公地址', '快递地址',
               '维系等级', '维系人', '备注', '状态']
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    level_colors = {'A': 'FFE0E0', 'B': 'FFF3E0', 'C': 'E3F2FD', 'D': 'F5F5F5'}
    for i, c in enumerate(contacts, 1):
        row = [i, c['category'], c['subcategory'] or '',
               c['unit_name'], c['department'], c['position'], c['name'],
               c['phone'], c['main_project'], c['office_address'], c['delivery_address'],
               c['maintenance_level'], c['maintenance_person'], c['remark'], c['status']]
        ws.append(row)
        for cell in ws[i + 1]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        # 维系等级着色（列索引现在是12）
        level_cell = ws.cell(row=i + 1, column=12)
        color = level_colors.get(c['maintenance_level'], 'FFFFFF')
        level_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    # 列宽
    widths = [6, 16, 12, 22, 14, 12, 12, 16, 28, 24, 24, 10, 12, 24, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # === Sheet2: 跟进记录 ===
    if contact_ids:
        placeholders = ','.join(['?'] * len(contact_ids))
        fq = f'''SELECT f.*, c.unit_name, c.name as contact_name, u.name as operator
                 FROM followups f
                 JOIN contacts c ON f.contact_id = c.id
                 LEFT JOIN users u ON f.created_by = u.id
                 WHERE f.contact_id IN ({placeholders})
                 ORDER BY f.follow_date DESC'''
        followups = db.execute(fq, contact_ids).fetchall()
    else:
        followups = []

    ws2 = wb.create_sheet('跟进记录')
    fu_headers = ['序号', '跟进日期', '跟进方式', '关联单位', '联系人', '跟进内容', '下次跟进日期', '操作人', '记录时间']
    ws2.append(fu_headers)

    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = PatternFill(start_color='4F6EF7', end_color='4F6EF7', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for i, f in enumerate(followups, 1):
        row = [i, f['follow_date'], f['follow_type'],
               f['unit_name'], f['contact_name'] or '',
               f['content'], f['next_follow_date'] or '',
               f['operator'] or '', f['created_at']]
        ws2.append(row)
        for cell in ws2[i + 1]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    fu_widths = [6, 14, 10, 22, 12, 50, 14, 12, 18]
    for i, w in enumerate(fu_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    log_action('导出数据', '', f'导出{len(contacts)}条联系人 + {len(followups)}条跟进记录')

    from urllib.parse import quote
    filename = f'营销人员信息统计表_{date.today().isoformat()}.xlsx'
    from flask import Response
    return Response(
        output.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ============================================================
# 模板下载
# ============================================================
@app.route('/import/template')
@login_required
def import_template():
    """下载导入模板Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
    from io import BytesIO

    wb = openpyxl.Workbook()

    # === Sheet1: 数据填写 ===
    ws = wb.active
    ws.title = '联系人数据'

    headers = [
        '大类', '子类', '单位名称', '部门', '职务', '姓名',
        '联系方式', '主要跟踪项目/目标', '办公地址', '快递地址',
        '维系等级', '维系人', '备注', '状态'
    ]
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F6EF7', end_color='4F6EF7', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    required_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        # 必填字段标注
        if headers[col_idx - 1] in ('大类', '单位名称'):
            cell.comment = Comment('此列为必填项', '系统')

    # 示例行（浅灰背景提示用户可删除）
    example_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
    examples = [
        ['业主单位', '政府部门', '湖北省发改委', '投资处', '处长', '张明',
         '13800138000', '武汉市轨道交通项目审批', '武汉市武昌区水果湖', '同办公地址',
         'A', '系统管理员', '省级发改主管部门', '活跃'],
        ['设计及咨询单位', '', '中交二航设计院', '市政所', '总工', '李华',
         '13900139000', '市政道路设计对接', '武汉市江汉区', '同办公地址',
         'B', '系统管理员', '重点跟踪设计单位', '活跃'],
        ['施工单位及监理', '', '中建三局', '工程部', '项目经理', '王强',
         '13700137000', '房建总包合作', '武汉市洪山区', '武汉市洪山区珞瑜路',
         'A', '系统管理员', '', '活跃'],
    ]
    for row_data in examples:
        ws.append(row_data)
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        # 标记示例行
        ws.cell(row=row_num, column=1).comment = Comment('示例数据，可删除后填写真实数据', '系统')

    # 数据验证 - 大类
    cat_list = ','.join(CATEGORIES.keys())
    dv_category = DataValidation(type='list', formula1=f'"{cat_list}"', allow_blank=False)
    dv_category.error = '请从下拉列表中选择有效大类'
    dv_category.errorTitle = '大类无效'
    dv_category.prompt = '点击下拉箭头选择大类'
    dv_category.promptTitle = '大类'
    ws.add_data_validation(dv_category)
    dv_category.add(f'A2:A1000')

    # 数据验证 - 子类（通过公式引用）
    # 由于openpyxl不支持级联验证，这里列出所有子类
    all_subs = []
    for subs in CATEGORIES.values():
        all_subs.extend(subs)
    if all_subs:
        sub_list = ','.join(all_subs)
        dv_sub = DataValidation(type='list', formula1=f'"{sub_list}"', allow_blank=True)
        dv_sub.error = '请从下拉列表中选择有效子类'
        dv_sub.errorTitle = '子类无效'
        dv_sub.prompt = '业主单位下的子类（其他大类可留空）'
        dv_sub.promptTitle = '子类'
        ws.add_data_validation(dv_sub)
        dv_sub.add(f'B2:B1000')

    # 数据验证 - 维系等级
    level_list = ','.join(MAINTENANCE_LEVELS.keys())
    dv_level = DataValidation(type='list', formula1=f'"{level_list}"', allow_blank=False)
    dv_level.error = '请输入 A / B / C / D'
    dv_level.errorTitle = '维系等级无效'
    dv_level.prompt = 'A=核心 B=重要 C=一般 D=潜在'
    dv_level.promptTitle = '维系等级'
    ws.add_data_validation(dv_level)
    dv_level.add(f'K2:K1000')

    # 数据验证 - 状态
    status_list = ','.join(CONTACT_STATUSES)
    dv_status = DataValidation(type='list', formula1=f'"{status_list}"', allow_blank=True)
    dv_status.error = '请输入：活跃 / 潜在 / 沉睡'
    dv_status.errorTitle = '状态无效'
    dv_status.prompt = '活跃 / 潜在 / 沉睡'
    dv_status.promptTitle = '客户状态'
    ws.add_data_validation(dv_status)
    dv_status.add(f'N2:N1000')

    # 列宽
    col_widths = [16, 14, 22, 14, 14, 12, 16, 28, 24, 24, 10, 14, 24, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = 'A2'

    # === Sheet2: 填写说明 ===
    ws2 = wb.create_sheet('填写说明')
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 60

    ws2['A1'] = '营销人员信息导入模板 - 填写说明'
    ws2['A1'].font = Font(bold=True, size=14, color='4F6EF7')
    ws2.merge_cells('A1:B1')

    instructions = [
        ('', ''),
        ('一、使用步骤', ''),
        ('步骤1', '下载本模板到本地'),
        ('步骤2', '在"联系人数据"工作表中填写客户信息（可删除示例行）'),
        ('步骤3', '保存文件后，回到网页点击"上传导入"'),
        ('步骤4', '确认导入结果，如有错误可修正后重新上传'),
        ('', ''),
        ('二、字段说明', ''),
        ('大类', '必填。下拉选择：业主单位 / 设计及咨询单位 / 施工单位及监理 / 供应商 / 运营方'),
        ('子类', '业主单位下的子类（政府部门/主管单位/投资方/学校），其他大类可留空'),
        ('单位名称', '必填。客户单位全称，如"湖北省发改委"'),
        ('部门', '部门名称，如"投资处"'),
        ('职务', '联系人职务，如"处长"'),
        ('姓名', '联系人姓名'),
        ('联系方式', '手机号或座机号'),
        ('主要跟踪项目/目标', '正在跟踪的工程项目或业务目标'),
        ('办公地址', '办公地点地址'),
        ('快递地址', '收件地址（可与办公地址相同）'),
        ('维系等级', '必填。下拉选择：A(核心) / B(重要) / C(一般) / D(潜在)'),
        ('维系人', '负责该客户的营销人员姓名'),
        ('备注', '其他补充信息'),
        ('状态', '下拉选择：活跃 / 潜在 / 沉睡（留空默认为"活跃"）'),
        ('', ''),
        ('三、注意事项', ''),
        ('注意1', '黄色底纹为必填字段（大类、单位名称）'),
        ('注意2', '浅蓝底纹行为示例数据，导入前请删除或替换'),
        ('注意3', '维系人请填写系统中已存在的用户姓名，否则留空'),
        ('注意4', '同一单位名称如已存在，将更新该单位信息而非重复新增'),
        ('注意5', '一次最多导入500条数据'),
    ]
    for row_idx, (key, desc) in enumerate(instructions, 2):
        ws2.cell(row=row_idx, column=1, value=key)
        ws2.cell(row=row_idx, column=2, value=desc)
        if key.startswith('一') or key.startswith('二') or key.startswith('三'):
            ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color='2C3E50')
        elif key.startswith('步骤') or key.startswith('注意'):
            ws2.cell(row=row_idx, column=1).font = Font(bold=True, color='4F6EF7')
        elif key:
            ws2.cell(row=row_idx, column=1).font = Font(bold=True)
            ws2.cell(row=row_idx, column=1).fill = PatternFill(
                start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    log_action('下载模板', '', '下载导入模板')

    from urllib.parse import quote
    from flask import Response
    filename = '营销人员信息导入模板.xlsx'
    return Response(
        output.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ============================================================
# 数据导入
# ============================================================
@app.route('/import/data', methods=['POST'])
@login_required
def import_data():
    """上传Excel文件并批量导入"""
    import openpyxl
    from io import BytesIO

    if 'file' not in request.files:
        return jsonify({'ok': False, 'msg': '请选择文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'ok': False, 'msg': '请选择文件'}), 400

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'msg': '仅支持 .xlsx / .xls 格式的Excel文件'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'Excel文件解析失败: {str(e)}'}), 400

    # 找到数据sheet
    ws = None
    for sheet_name in ['联系人数据', 'Sheet1', '营销人员信息统计表']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break
    if not ws:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({'ok': False, 'msg': '文件中没有数据行（仅含表头）'}), 400

    # 表头映射 - 智能匹配列名
    header_row = [str(h).strip() if h else '' for h in rows[0]]
    col_map = {}
    expected_fields = {
        '大类': ['大类', '分类', 'category'],
        '子类': ['子类', '小类', 'subcategory'],
        '单位名称': ['单位名称', '单位', 'unit_name'],
        '部门': ['部门', 'department'],
        '职务': ['职务', 'position'],
        '姓名': ['姓名', 'name'],
        '联系方式': ['联系方式', '电话', '手机', 'phone'],
        '主要跟踪项目/目标': ['主要跟踪项目/目标', '主要跟踪项目、目标', '主要跟踪项目', '跟踪项目', 'main_project'],
        '办公地址': ['办公地址', '办公地址（办公室）', 'office_address'],
        '快递地址': ['快递地址', 'delivery_address'],
        '维系等级': ['维系等级', 'maintenance_level'],
        '维系人': ['维系人', 'maintenance_person'],
        '备注': ['备注', 'remark'],
        '状态': ['状态', 'status'],
    }
    for field, aliases in expected_fields.items():
        for idx, header in enumerate(header_row):
            if header in aliases:
                col_map[field] = idx
                break

    # 必须有单位名称列
    if '单位名称' not in col_map:
        return jsonify({'ok': False, 'msg': '未找到"单位名称"列，请使用标准模板或确保表头包含"单位名称"'}), 400

    db = get_db()
    success_count = 0
    update_count = 0
    error_count = 0
    errors = []
    valid_categories = set(CATEGORIES.keys())
    valid_levels = set(MAINTENANCE_LEVELS.keys())
    valid_statuses = set(CONTACT_STATUSES)

    # 获取已有单位列表（用于判重）
    existing_units = {}
    existing_rows = db.execute('SELECT id, unit_name FROM contacts').fetchall()
    for r in existing_rows:
        existing_units[r['unit_name']] = r['id']

    data_rows = rows[1:]  # 去掉表头
    total_rows = len(data_rows)

    for row_idx, row in enumerate(data_rows, 2):  # Excel行号从2开始
        # 跳过完全空行
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue

        def get_cell(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return ''
            val = row[idx]
            return str(val).strip() if val is not None else ''

        unit_name = get_cell('单位名称')
        category = get_cell('大类')

        # 校验必填
        if not unit_name:
            errors.append(f'第{row_idx}行：单位名称为空，已跳过')
            error_count += 1
            continue
        if not category:
            errors.append(f'第{row_idx}行：大类为空，已跳过')
            error_count += 1
            continue
        if category not in valid_categories:
            errors.append(f'第{row_idx}行：大类"{category}"无效，应为：{"/".join(valid_categories)}')
            error_count += 1
            continue

        # 校验维系等级
        level = get_cell('维系等级') or 'B'
        if level not in valid_levels:
            errors.append(f'第{row_idx}行：维系等级"{level}"无效，应为A/B/C/D，已默认设为B')
            level = 'B'

        # 校验状态
        status = get_cell('状态') or '活跃'
        if status not in valid_statuses:
            errors.append(f'第{row_idx}行：状态"{status}"无效，应为活跃/潜在/沉睡，已默认设为活跃')
            status = '活跃'

        data = {
            'category': category,
            'subcategory': get_cell('子类'),
            'unit_name': unit_name,
            'department': get_cell('部门'),
            'position': get_cell('职务'),
            'name': get_cell('姓名'),
            'phone': get_cell('联系方式'),
            'main_project': get_cell('主要跟踪项目/目标'),
            'office_address': get_cell('办公地址'),
            'delivery_address': get_cell('快递地址'),
            'maintenance_level': level,
            'maintenance_person': get_cell('维系人'),
            'remark': get_cell('备注'),
            'status': status,
        }

        try:
            if unit_name in existing_units:
                # 更新已有记录
                cid = existing_units[unit_name]
                db.execute('''UPDATE contacts SET
                    category=?, subcategory=?, unit_name=?, department=?, position=?, name=?,
                    phone=?, main_project=?, office_address=?, delivery_address=?,
                    maintenance_level=?, maintenance_person=?, remark=?, status=?,
                    updated_at=datetime('localtime') WHERE id=?''',
                    (*data.values(), cid))
                update_count += 1
            else:
                # 新增
                db.execute('''INSERT INTO contacts
                    (category, subcategory, unit_name, department, position, name, phone,
                     main_project, office_address, delivery_address, maintenance_level,
                     maintenance_person, remark, status, owner_id, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (*data.values(), session['user_id'], session['user_id']))
                new_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
                existing_units[unit_name] = new_id
                success_count += 1
        except Exception as e:
            errors.append(f'第{row_idx}行：数据库写入失败 - {str(e)}')
            error_count += 1

    db.commit()
    log_action('导入数据', '', f'导入: 新增{success_count}条, 更新{update_count}条, 失败{error_count}条')

    # 限制错误信息数量
    if len(errors) > 20:
        errors = errors[:20] + [f'... 共{len(errors)}条错误信息']

    return jsonify({
        'ok': True,
        'total': total_rows,
        'success': success_count,
        'updated': update_count,
        'errors': error_count,
        'error_details': errors,
    })


# ============================================================
# API - 获取子类
# ============================================================
@app.route('/api/subcategories')
@login_required
def api_subcategories():
    category = request.args.get('category', '')
    subs = CATEGORIES.get(category, [])
    return jsonify({'subcategories': subs})


# ============================================================
# 导入页面
# ============================================================
@app.route('/import')
@login_required
def import_page():
    """数据导入页面"""
    db = get_db()
    total = db.execute('SELECT COUNT(*) as c FROM contacts').fetchone()['c']
    return render_template('import.html', total=total)


# ============================================================
# 错误处理
# ============================================================
@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', code=403, msg='权限不足，请联系管理员'), 403


@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404, msg='页面不存在'), 404


@app.errorhandler(500)
def server_error(e):
    return render_template('error.html', code=500, msg='服务器内部错误'), 500


# ============================================================
# 模板上下文
# ============================================================
@app.context_processor
def inject_globals():
    return {
        'current_user_name': session.get('name', ''),
        'current_username': session.get('username', ''),
        'is_admin': session.get('role') == 'admin',
        'current_access_level': session.get('access_level', 'L1'),
        'current_team': session.get('team', ''),
        'categories': CATEGORIES,
        'levels': MAINTENANCE_LEVELS,
        'access_levels': ACCESS_LEVELS,
        'now': datetime.now(),
    }


# ============================================================
# 自动初始化数据库（模块加载时执行，兼容 gunicorn 等 WSGI 服务器）
# ============================================================
init_db()

# ============================================================
# 启动
# ============================================================
if __name__ == '__main__':
    port = app.config['PORT']
    print('=' * 60)
    print('  营销人员信息管理系统')
    print(f'  访问地址: http://0.0.0.0:{port}')
    print('  管理员账号: admin / admin123')
    print('  营销员账号: sales01 / sales123')
    print('=' * 60)
    app.run(host='0.0.0.0', port=port, debug=app.config['DEBUG'])
